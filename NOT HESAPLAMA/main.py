import openpyxl
from openpyxl.reader.excel import load_workbook

wb = load_workbook("veri.xlsx")
ws = wb.active

#Tüm Verileri Getiren Kodx
'''
for satir in range(1,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()
'''
no=[]
ogrno=[]
adi=[]
soyadi=[]
sinif=[]
atipi=[]
viz=[]
vizmaz=[]
final=[]
but=[]
for row in ws.iter_rows(min_row=1, min_col=1, max_row=73, max_col=1):
    for cell in row:
        no.append(cell.value)
no.pop(0)
for row in ws.iter_rows(min_row=1, min_col=2, max_row=73, max_col=2):
    for cell in row:
        ogrno.append(cell.value)
ogrno.pop(0)
for row in ws.iter_rows(min_row=1, min_col=3, max_row=73, max_col=3):
    for cell in row:
        adi.append(cell.value)
adi.pop(0)
for row in ws.iter_rows(min_row=1, min_col=4, max_row=73, max_col=4):
    for cell in row:
        soyadi.append(cell.value)
soyadi.pop(0)
for row in ws.iter_rows(min_row=1, min_col=5, max_row=73, max_col=5):
    for cell in row:
        sinif.append(cell.value)
sinif.pop(0)
for row in ws.iter_rows(min_row=1, min_col=6, max_row=73, max_col=6):
    for cell in row:
        atipi.append(cell.value)
atipi.pop(0)

for row in ws.iter_rows(min_row=1, min_col=11, max_row=73, max_col=11):
    for cell in row:
        viz.append(cell.value)
viz.pop(0)
for row in ws.iter_rows(min_row=1, min_col=12, max_row=73, max_col=12):
    for cell in row:
        vizmaz.append(cell.value)
vizmaz.pop(0)
for row in ws.iter_rows(min_row=1, min_col=13, max_row=73, max_col=13):
    for cell in row:
        final.append(cell.value)
final.pop(0)
for row in ws.iter_rows(min_row=1, min_col=14, max_row=73, max_col=14):
    for cell in row:
        but.append(cell.value)
but.pop(0)
vizeler=[]
ogrnotlar=[]
for row in ws.iter_rows(min_row=1, min_col=11, max_row=73, max_col=14):
    for cell in row:
        if cell.value==None:
            vizeler.append(0)
        else:
            vizeler.append(cell.value)
for x in range(4):
    vizeler.pop(0)
#print(vizeler) #excel ilk dört veri temizliği  
z=0
f=0
while z<72:
    a=[vizeler[f],vizeler[f+1],vizeler[f+2],vizeler[f+3]]
    ogrnotlar.append(a)
    z=z+1
    f=f+4
#print(ogrnotlar) #ÖĞRENCİLERİN VİZE/VİZEMAZERET VE FİNAL/BÜTÜNLEME NOTLARINI YAZDIRABİLİYORUZ.
sayac=0
x=0
y=0
so=0
hbn=[]
for x in range(72): #hocam buraya len(vizenotlar) yazıyorum sürekli hata alıyordum o yüzden range(72) yazarak düzeltebildim.)
    ortalama=((ogrnotlar[y][0]*0.4)+(ogrnotlar[y][1]*0.4)+(ogrnotlar[y][2]*0.6))
    if ortalama==0:
        sayac=sayac+1 #sınava hiç girmeyen kaç kişi var onun tespiti.
    #else bloğu ile başka bir sayaç açarak sınava giren öğrenci saysını bulabiliriz fakat döngünün kaç kez döndüğünü bildiğimiz için 72-sayac yaparak sınava giren öğrenci sayısını buluyoruz.
    ortalama=round(ortalama)
    so=so+ortalama
    hbn.append(ortalama)
    y=y+1
#print(hbn) #ÖĞRENCİLERİN HBN'Sİ.
sgos=72-sayac 
#print(sgos) #sınava giren öğrenci sayısı

so=so/sgos #sınava hiç girmeyen öğrencileri ortalamaya dahil etmiyoruz.
#print(so) #SINIF ORTALAMASINI BULUYORUZ.
'''
liste2 = []
varyans = 0
for sayi in range(72):
	varyans += (hbn[sayi] - so)**2
#print(varyans) #varyans
stdsapma = varyans ** (1/2)
#print(stdsapma) #standart sapma
'''
varyans=0
for sayi in hbn:
	varyans += (sayi - so)**2
varyans=varyans/sgos
stdsapma = varyans ** (1/2)
#print(stdsapma)

zdegerleri = []
for ort in hbn:
	zdegerleri.append((ort - so)/stdsapma)
#print(zdegerleri) #zdeğerlerin

tdegerleri = []
for zdegeri in zdegerleri:
	tdegerleri.append(zdegeri*10+50)
#print(tdegerleri) #tdeğerleri

harfnotları=[]

if so>=70:
    for x in hbn:
        if x>=90:
            harfnotları.append("AA")
        if x>=85 and x<=89:
            harfnotları.append("BA")
        if x>=80 and x<=84:
            harfnotları.append("BB")
        if x>=75 and x<=79:
            harfnotları.append("CB")
        if x>=70 and x<=74:
            harfnotları.append("CC")
        if x>=65 and x<=69:
            harfnotları.append("DC")
        if x>=60 and x<=64:
            harfnotları.append("DD")
        if x>=50 and x<=59:
            harfnotları.append("FD")
        if x<=49:
            harfnotları.append("FF")
        
if sgos<30:
    aa=int(input("AA notu sınırı belirleyin (ALT SINIR)"))
    ba=int(input("BA notu sınırı belirleyin (ALT SINIR)"))
    bb=int(input("BB notu sınırı belirleyin (ALT SINIR)"))
    cb=int(input("CB notu sınırı belirleyin (ALT SINIR)"))
    cc=int(input("CC notu sınırı belirleyin (ALT SINIR)"))
    dc=int(input("DC notu sınırı belirleyin (ALT SINIR)"))
    dd=int(input("DD notu sınırı belirleyin (ALT SINIR)"))
    fd=int(input("FD notu sınırı belirleyin (ALT SINIR)"))
    ff=int(input("FF notu sınırı belirleyin (ALT SINIR)"))
    for x in hbn:
        if x>=aa:
            harfnotları.append("AA")
        if x>=ba and x<aa:
            harfnotları.append("BA")
        if x>=bb and x<ba:
            harfnotları.append("BB")
        if x>=cb and x<bb:
            harfnotları.append("CB")
        if x>=cc and x<cb:
            harfnotları.append("CC")
        if x>=dc and x<cc:
            harfnotları.append("DC")
        if x>=dd and x<dc:
            harfnotları.append("DD")
        if x>=fd and x<dd:
            harfnotları.append("FD")
        if x<=ff:
            harfnotları.append("FF")


if sgos>=30:
    if so>=62.5:
        for x in tdegerleri:
            if x>=61:
                harfnotları.append("AA")
            if x>=56.99 and x<=60.99:
                harfnotları.append("BA")
            if x>=51 and x<=55.99:
                harfnotları.append("BB")
            if x>=46 and x<=50.99:
                harfnotları.append("CB")
            if x>=41.99 and x<=45.99:
                harfnotları.append("CC")
            if x>=36 and x<=40.99:
                harfnotları.append("DC")
            if x>=31.99 and x<=35.99:
                harfnotları.append("DD")
            if x>=26 and x<=30.99:
                harfnotları.append("FD")
            if x<=26:
                harfnotları.append("FF")
    if so>=57.50 and so<=62.49:
        for x in tdegerleri:
            if x>=63:
                harfnotları.append("AA")
            if x>=58 and x<=62.99:
                harfnotları.append("BA")
            if x>=53 and x<=57.99:
                harfnotları.append("BB")                
            if x>=48 and x<=52.99:
                harfnotları.append("CB")
            if x>=43 and x<=47.99:
                harfnotları.append("CC")
            if x>=38 and x<=42.99:
                harfnotları.append("DC")
            if x>=33 and x<=37.99:
                harfnotları.append("DD")
            if x>=28 and x<=32.99:
                harfnotları.append("FD")
            if x<=28:
                harfnotları.append("FF")
    if so>=52.50 and so<=57.49:
        for x in tdegerleri:
            if x>=65:
                harfnotları.append("AA")
            if x>=60 and x<=64.99:
                harfnotları.append("BA")
            if x>=55 and x<=59.99:
                harfnotları.append("BB")
            if x>=50 and x<=54.99:
                harfnotları.append("CB")
            if x>=45 and x<=49.99:
                harfnotları.append("CC")
            if x>=40 and x<=44.99:
                harfnotları.append("DC")
            if x>=35 and x<=39.99:
                harfnotları.append("DD")
            if x>=30 and x<=34.99:
                harfnotları.append("FD")
            if x<=30:
                harfnotları.append("FF")
    if so>=47.50 and so<=52.49:
        for x in tdegerleri:
            if x>=67:
                harfnotları.append("AA")
            if x>=62 and x<=66.99:
                harfnotları.append("BA")
            if x>=57 and x<=61.99:
                harfnotları.append("BB")
            if x>=52 and x<=56.99:
                harfnotları.append("CB")
            if x>=47 and x<=51.99:
                harfnotları.append("CC")
            if x>=42 and x<=46.99:
                harfnotları.append("DC")
            if x>=37 and x<=41.99:
                harfnotları.append("DD")
            if x>=32 and x<=36.99:
                harfnotları.append("FD")
            if x<=32:
                harfnotları.append("FF")
    if so>=42.50 and so<=47.49:
        for x in tdegerleri:
            if x>=69:
                harfnotları.append("AA")
            if x>=64 and x<=68.99:
                harfnotları.append("BA")
            if x>=59 and x<=63.99:
                harfnotları.append("BB")
            if x>=54 and x<=58.99:
                harfnotları.append("CB")
            if x>=49 and x<=53.99:
                harfnotları.append("CC")
            if x>=44 and x<=48.99:
                harfnotları.append("DC")
            if x>=39 and x<=43.99:
                harfnotları.append("DD")
            if x>=34 and x<=38.99:
                harfnotları.append("FD")
            if x<=34:
                harfnotları.append("FF")
    if so>=0 and so<=42.49:
        for x in tdegerleri:
            if x>=71:
                harfnotları.append("AA")
            if x>=66 and x<=70.99:
                harfnotları.append("BA")
            if x>=61 and x<=65.99:
                harfnotları.append("BB")
            if x>=56 and x<=60.99:
                harfnotları.append("CB")
            if x>=51 and x<=55.99:
                harfnotları.append("CC")
            if x>=46 and x<=50.99:
                harfnotları.append("DC")
            if x>=41 and x<=45.99:
                harfnotları.append("DD")
            if x>=36 and x<=40.99:
                harfnotları.append("FD")
            if x<=36:
                harfnotları.append("FF")

gecme=[]
for x in harfnotları:
    if x=="FF":
        gecme.append("KALDI")
    else:
        gecme.append("GEÇTİ")

#print(gecme)
#tskor zdeğer veya standart sapmadaki yanlışlıktan dolayı yanlış sonuç çıkıyor.

wb = load_workbook("sonuc.xlsx")
ws = wb.active
#ws.append(["NO","ÖĞRENCİ NO","ADI","SOYADI","SINIF","A.TİPİ","HBN","T SKOR","HARF NOTLARI","GEÇME","VİZE","VİZE MAZARET","FİNAL","BÜTÜNLEME"])
for x in range(72):
    ws.append([no[x],ogrno[x],adi[x],soyadi[x],sinif[x],atipi[x],hbn[x],tdegerleri[x],harfnotları[x],gecme[x],viz[x],vizmaz[x],final[x],but[x]])


wb.save("sonuc.xlsx")

'''
Excel'in belirli satırına yazdırmak
ws['A1'] = 42
ws['B3'] = "Merhaba"
ws.append([1, 2, 3, "Görkem", "Görür"])
'''

